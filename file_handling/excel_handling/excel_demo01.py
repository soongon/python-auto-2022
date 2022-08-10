import openpyxl
import datetime

# 새로운 엑셀 파일을 만들어서 데이터를 셀에 쓰기
wb = openpyxl.workbook.Workbook()

ws = wb.active
ws['A1'] = 'Hello world'
ws['B3'] = 35
ws['D4'] = 57.34
ws['E5'] = datetime.datetime.now()
ws.append(['red', 'blue', 'white'])

wb.save('sample.xlsx')

